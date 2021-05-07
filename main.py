# -*- coding: utf-8 -*-

"""
Created on Wed Feb  3 13:17:34 2021

@author: brian

備註：PMS裡面是right補給left

"""

import cv2
import numpy as np
from openpyxl import Workbook
from metrics import SSIM
import utils
import occlusion_utils

if __name__ == '__main__':
    np.set_printoptions(suppress=True)
    #%% 計算中心圖抽樣放大前後誤差diff(為int32有正有負)
    print("Calculating difference...")
    mid = cv2.imread("warping_dataset/bike/HD_data/8.png")#讀取中心圖
    height = mid.shape[0]
    width = mid.shape[1]
    scale_percent = 1/4 # percent of original size #換圖要改
    mid_down = cv2.resize(mid, (int(height*scale_percent), int(width*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
    mid_re = cv2.resize(mid_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
    dif = (mid.astype(np.int32) - mid_re.astype(np.int32))[:,:,:].copy()
    #%% 將圖放大縮小後儲存
    """
    for i in range(17,21):
            right = cv2.imread("warping_dataset/bike/HD_data/8.png")#換圖要改
            right_down = cv2.resize(right, (int(width*scale_percent), int(height*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
            right_re = cv2.resize(right_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
            left = cv2.imread("warping_dataset/bike/HD_data/"+ str(i) +".png")#換圖要改
            left_down = cv2.resize(left, (int(height*scale_percent), int(width*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
            left_re = cv2.resize(left_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
            cv2.imwrite("warping_dataset/bike/LR_data_pair/" + str(i) + "_right.png", right_re)
            cv2.imwrite("warping_dataset/bike/LR_data_pair/" + str(i) + "_left.png", left_re)
    
    for i in range(17):
            right = cv2.imread("warping_dataset/bike/HD_data/8.png")#換圖要改
            right_down = cv2.resize(right, (int(width*scale_percent), int(height*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
            right_re = cv2.resize(right_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
            left = cv2.imread("warping_dataset/bike/HD_data/"+ str(i) +".png")#換圖要改
            left_down = cv2.resize(left, (int(height*scale_percent), int(width*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
            left_re = cv2.resize(left_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
            cv2.imwrite("warping_dataset/bike/LR_data_pair/" + str(i) + "_right.png", right_re)
            cv2.imwrite("warping_dataset/bike/LR_data_pair/" + str(i) + "_left.png", left_re)
    """
#%% 把每一個disparity map還原回真實視差值 [0,255]->[]
    disp_arr_sub, disp_arr_mid = utils.recoverDisValue(height, width)
    # =============================================================================
    #     #%% 計算disparity前二大跟前二小的值屬於哪張disparity map
    #     print("Calculating MAX and min disparity values...")
    #     utils.disp_max_min(disp_arr_mid, height, width, 17)
    #     
    #     #%% 看disparity map值的分布狀態
    #     print("Check diparity distribution...")
    #     utils.disp_distribution(disp_arr_mid, height, width)
    # =============================================================================
    #%% 算disparity error越大，是否實際pixel值也差越多
    #utils.disp_error_pix_error(disp_arr_mid, disp_arr_sub, height, width)
    #%% 處理occlusion區域 
    unwarpedMap = "inde" #one, ring, inde, fusion
    disp_arr_sub = disp_arr_sub.astype(np.int32)
    disp_arr_mid = disp_arr_mid.astype(np.int32)
    print("Occlusion Handling by "+ unwarpedMap +"...")
    threshold = 1
    occlusion_arr = np.zeros((17, height, width), dtype = np.uint8)
    for k in range(17):
        if( k >= 0 and k <= 3):
            for i in range(height):
                for j in range(width):
                    if((i + disp_arr_mid[k,i,j]) < height):
                        if(abs(disp_arr_mid[k,i,j] - disp_arr_sub[k,i + disp_arr_mid[k,i,j],j]) > threshold):
                            occlusion_arr[k,i,j] = 1
                    else:
                        occlusion_arr[k,i,j] = 1
                        
        elif( k >= 4 and k <= 7):
            for i in range(height):
                for j in range(width):
                    if((j + disp_arr_mid[k,i,j]) < width):
                        if(abs(disp_arr_mid[k,i,j] - disp_arr_sub[k,i,j + disp_arr_mid[k,i,j]]) > threshold):
                            occlusion_arr[k,i,j] = 1
                    else:
                        occlusion_arr[k,i,j] = 1
                        
        elif( k >= 9 and k <= 12):
            for i in range(height):
                for j in range(width):
                    if((j - disp_arr_mid[k,i,j]) >= 0):
                        if(abs(disp_arr_mid[k,i,j] - disp_arr_sub[k,i,j - disp_arr_mid[k,i,j]]) > threshold):
                            occlusion_arr[k,i,j] = 1
                    else:
                        occlusion_arr[k,i,j] = 1
                        
        elif( k >= 13 and k <= 16):
            for i in range(height):
                for j in range(width):
                    if((i - disp_arr_mid[k,i,j]) >= 0):
                        if(abs(disp_arr_mid[k,i,j] - disp_arr_sub[k,i - disp_arr_mid[k,i,j],j]) > threshold):
                            occlusion_arr[k,i,j] = 1
                    else:
                        occlusion_arr[k,i,j] = 1
                            
    if(unwarpedMap == "one"):
        disp_arr_mid, one_occlusion = occlusion_utils.oneOcc_dilateFirst(occlusion_arr, disp_arr_mid, height, width)
    elif(unwarpedMap == "ring"):
        disp_arr_mid, ring_occlusion = occlusion_utils.ringOcc_dilateFirst(occlusion_arr, disp_arr_mid, height, width)
    elif(unwarpedMap == "inde"):
        disp_arr_mid, inde_occlusion = occlusion_utils.indeOcc_dilateFirst(occlusion_arr, disp_arr_mid, height, width)

    
                    
    #%%選擇如何處理outliers, output:disp
    
    #測試圖片都是用十字16張圖片，之後有需要再從utils裡面改
    outlier_case = 0
    
    #用median來處理outlier...0
    if(outlier_case == 0):
        print("Remove outliers by median...")
        disp = utils.outlier_median(disp_arr_mid, height, width)
        method = "median"
        
    #用盒鬚圖來處理outlier...1
    if(outlier_case == 1):
        print("Remove outliers by box plot...")
        disp = utils.outlier_boxplot(disp_arr_mid, height, width)
        method = "boxplot"
        
    #用標準差來去除outlier...2
    if(outlier_case == 2):
        print("Remove outliers by deviation...")
        disp = utils.outlier_dev(disp_arr_mid, height, width, 1)
        method = "stdev"
    
    #將disparity maps在每個pixel中最大跟最小各兩個值刪除後平均(論文方法)...3
    if(outlier_case == 3):
        print("Remove outliers by 去頭去尾...")
        disp = utils.outlier_paper(disp_arr_mid, height, width)
        method = "blend"
    
    #%% 做基於blended disparity map的warpping
    print("Warping by " + method + " disparity map...")
    wb = Workbook()
    ws = wb.active
    ws.append(["","bilinear_psnr","compensate_"+ method +"_psnr","bilinear_SSIM","compensate_"+ method +"_SSIM"])
    diftmp = dif.astype(np.float32)
    for k in range(17):
        
        if(k!=8):
            print("Warping difference map by "+ method +" disparity on " + str(k) + "...")
            sub = cv2.imread("warping_dataset/bike/HD_data/"+ str(k) +".png")
            sub_down = cv2.resize(sub, (int(width*scale_percent), int(height*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
            sub_re = cv2.resize(sub_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
            difUp = cv2.resize(diftmp, (int(width*4), int(height*4)), interpolation=cv2.INTER_LINEAR)
            dispUp = cv2.resize(disp, (int(width*4), int(height*4)), interpolation=cv2.INTER_LINEAR)
            
            occUp = np.zeros((17, int(width*4), int(height*4)))
            for i in range(17):
                occUp[i] = cv2.resize(inde_occlusion[i], (int(width*4), int(height*4)), interpolation=cv2.INTER_LINEAR)
# =============================================================================
#             occUp = np.zeros((4,int(width*4), int(height*4)))
#             for i in range(4):
#                 occUp[i] = cv2.resize(ring_occlusion[i], (int(width*4), int(height*4)), interpolation=cv2.INTER_LINEAR)
# =============================================================================
            #occUp = cv2.resize(one_occlusion, (int(width*4), int(height*4)), interpolation=cv2.INTER_LINEAR)
            
            warpUp = np.zeros((int(width*4), int(height*4), 3))
            new_sub = (sub_re.copy()).astype(np.int32)

            if(k == 0 or k == 4 or k == 12 or k == 16):
                tmp = 0
            elif(k == 1 or k == 5 or k == 11 or k == 15):
                tmp = 1
            elif(k == 2 or k == 6 or k == 10 or k == 14):
                tmp = 2
            elif(k == 3 or k == 7 or k == 9 or k == 13):
                tmp = 3  
            
# =============================================================================
#             #計算all occlusion的點warp後會到哪裡
#             newOcc = np.zeros((int(width), int(height)))
#             if( k >= 0 and k <= 3):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i + disp[i,j]) < height and one_occlusion[i,j] == 1):
#                             newOcc[i + disp[i,j], j] = 1
#                             
#             elif( k >= 4 and k <= 7):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j + disp[i,j]) < width and one_occlusion[i,j] == 1):
#                             newOcc[i, j + disp[i,j]] = 1
#                             
#             elif( k >= 9 and k <= 12):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j - disp[i,j]) >= 0 and one_occlusion[i,j] == 1):
#                             newOcc[i, j - disp[i,j]] = 1
#                             
#             elif( k >= 13 and k <= 16):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i - disp[i,j]) >= 0 and one_occlusion[i,j] == 1):
#                             newOcc[i - disp[i,j], j] = 1               
# =============================================================================
                
            #disparity map跟difference map都放大4被後做warping, 且都為occlusion的部分不補
            if(k >= 0 and k <= 3):
                for i in range(height*4):
                    for j in range(width*4):
                        if((i + dispUp[i, j]) < height*4 and occUp[k,i,j] != 1):
                            warpUp[i + dispUp[i, j], j, :] = difUp[i, j, :]

            elif(k >= 4 and k <= 7):
                for i in range(height*4):
                    for j in range(width*4):
                        if((j + dispUp[i, j]) < width*4 and occUp[k,i,j] != 1):
                            warpUp[i, j + dispUp[i, j], :] = difUp[i, j, :]

            elif(k >= 9 and k <= 12):
                for i in range(height*4):
                    for j in range(width*4):
                        if((j - dispUp[i, j]) >= 0 and occUp[k,i,j] != 1):
                            warpUp[i, j - dispUp[i, j], :] = difUp[i, j, :]

            elif(k >= 13 and k <= 16):
                for i in range(height*4):
                    for j in range(width*4):
                        if((i - dispUp[i, j]) >= 0 and occUp[k,i,j] != 1):
                            warpUp[i - dispUp[i, j], j, :] = difUp[i, j, :]

            warpDown = cv2.resize(warpUp, (width, height),interpolation=cv2.INTER_LINEAR)
            warped = warpDown.astype(np.int32)   
            for i in range(height):
                for j in range(width):
                    new_sub[i,j] += warped[i,j]
            #warped = warped.astype(np.uint8)
# =============================================================================
#             cv2.imshow("a", warped)
#             cv2.waitKey(0)
# =============================================================================
                 
            
# =============================================================================
#             #看occlusion位置
#             for i in range(height):
#                 for j in range(width):
#                     if(one_occlusion[i,j]==1):
#                         one_occlusion[i,j] = 255
#             one_occlusion = one_occlusion.astype(np.uint8)
#             cv2.imshow("a", one_occlusion)
#             cv2.imwrite("occluded_difmap.png", warped)
#             cv2.imwrite("occlusion_map.png", one_occlusion)
# =============================================================================
      
#沒有放大四倍再warp
# =============================================================================
#             if( k >= 0 and k <= 3):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i + disp[i,j]) < height):
#                             new_sub[i + disp[i,j], j] += dif[i,j]
#                             
#             elif( k >= 4 and k <= 7):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j + disp[i,j]) < width):
#                             new_sub[i, j + disp[i,j]] += dif[i,j]
#                             
#             elif( k >= 9 and k <= 12):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j - disp[i,j]) >= 0):
#                             new_sub[i, j - disp[i,j]] += dif[i,j]
#                             
#             elif( k >= 13 and k <= 16):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i - disp[i,j]) >= 0):
#                             new_sub[i - disp[i,j], j] += dif[i,j]                
# =============================================================================
            
            new_sub = np.clip(new_sub, 0, 255)
            new_sub = new_sub.astype(np.uint8)
            
            SSIM_R = SSIM(sub[:,:,2], sub_re[:,:,2])
            SSIM_G = SSIM(sub[:,:,1], sub_re[:,:,1])
            SSIM_B = SSIM(sub[:,:,0], sub_re[:,:,0])
            SSIM_sub_re = (SSIM_R + SSIM_B + SSIM_G)/3
             
            SSIM_R = SSIM(sub[:,:,2], new_sub[:,:,2])
            SSIM_G = SSIM(sub[:,:,1], new_sub[:,:,1])
            SSIM_B = SSIM(sub[:,:,0], new_sub[:,:,0])
            SSIM_new_sub = (SSIM_R + SSIM_B + SSIM_G)/3
            
            ws.append([str(k), cv2.PSNR(sub, sub_re), cv2.PSNR(sub, new_sub), SSIM_sub_re, SSIM_new_sub])
            print(str(k))
            print("sub_re PSNR: " + str(cv2.PSNR(sub, sub_re)))
            print("new_sub PSNR: " + str(cv2.PSNR(sub, new_sub)))
            print("SSIM sub_re: " + str(SSIM_sub_re))
            print("SSIM new_sub: " + str(SSIM_new_sub))
            #cv2.imwrite("warping_dataset/bike/after_compensate/"+ str(k) + "_" + method +"_withoutOcclusion_twoside_thresh1.png", new_sub) 
    wb.save('warping_dataset/bike/testing.xlsx')
     
# =============================================================================
#     #%% 做基於own disparity map 的warping
#     print("Warping by own disparity map...")
#     wb = Workbook()
#     ws = wb.active
#     ws.append(["","bilinear_psnr","compensate_own_psnr","bilinear_SSIM","compensate_own_SSIM"])
#     
#     for k in range(17):
#         if(k!=8):
#             print("Warping difference map by own disparity map on " + str(k) + "...")
#             sub = cv2.imread("warping_dataset/bike/HD_data/"+ str(k) +".png")
#             sub_down = cv2.resize(sub, (int(width*scale_percent), int(height*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
#             sub_re = cv2.resize(sub_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
#             new_sub = sub_re.astype(np.int32)
#             #------改blend或own disparity改這邊------
#             if( k >= 0 and k <= 3):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i + disp_arr_mid[k,i,j]) < height):
#                             new_sub[i + disp_arr_mid[k,i,j], j] += dif[i,j]
#                             
#             elif(k >= 4 and k <= 7):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j + disp_arr_mid[k,i,j]) < width):
#                             new_sub[i, j + disp_arr_mid[k,i,j]] += dif[i,j]
#                             
#             elif( k >= 9 and k <= 12):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j - disp_arr_mid[k,i,j]) > 0):
#                             new_sub[i, j - disp_arr_mid[k,i,j]] += dif[i,j]
#                             
#             elif( k >= 13 and k <= 16):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i - disp_arr_mid[k,i,j]) > 0):
#                             new_sub[i - disp_arr_mid[k,i,j], j] += dif[i,j]                
#             
#             new_sub = np.clip(new_sub, 0, 255)
#             new_sub = new_sub.astype(np.uint8)
#             
#             SSIM_R = SSIM(sub[:,:,2], sub_re[:,:,2])
#             SSIM_G = SSIM(sub[:,:,1], sub_re[:,:,1])
#             SSIM_B = SSIM(sub[:,:,0], sub_re[:,:,0])
#             SSIM_sub_re = (SSIM_R+SSIM_B+SSIM_G)/3
#             
#             
#             SSIM_R = SSIM(sub[:,:,2], new_sub[:,:,2])
#             SSIM_G = SSIM(sub[:,:,1], new_sub[:,:,1])
#             SSIM_B = SSIM(sub[:,:,0], new_sub[:,:,0])
#             SSIM_new_sub = (SSIM_R+SSIM_B+SSIM_G)/3
#             
#             ws.append([str(k), cv2.PSNR(sub,sub_re), cv2.PSNR(sub, new_sub), SSIM_sub_re, SSIM_new_sub])
#             print(str(k))
#             print("sub_re PSNR: " + str(cv2.PSNR(sub, sub_re)))
#             print("new_sub PSNR: " + str(cv2.PSNR(sub, new_sub)))
#             print("SSIM sub_re: " + str(SSIM_sub_re))
#             print("SSIM new_sub: " + str(SSIM_new_sub))
#             cv2.imwrite("warping_dataset/bike/after_compensate/"+ str(k) +"_own_withoutOcclusion_twoside_thresh1.png", new_sub) 
#     
#     wb.save('warping_dataset/bike/own_compensate.xlsx')   
#     #%% 以中心點為軸心對稱disparity map做平均的warping
#     print("Warping by blended symmetric disparity map...")
#     wb = Workbook()
#     ws = wb.active
#     ws.append(["","bilinear_psnr","compensate_symmetric_psnr","bilinear_SSIM","compensate_symmetric_SSIM"])
#     
#     for k in range(17):
#         if(k!=8):
#             sym = 16-k
#             disp_sym = (disp_arr_mid[k] + disp_arr_mid[sym])/2
#             disp_sym = disp_sym.astype(np.uint8)
#     
#             print("Warping difference by blended symmetric disparity map on " + str(k) + "...")
#             sub = cv2.imread("warping_dataset/bike/HD_data/"+ str(k) +".png")
#             sub_down = cv2.resize(sub, (int(width*scale_percent), int(height*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
#             sub_re = cv2.resize(sub_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
#             new_sub = sub_re.astype(np.int32)
#             if( k >= 0 and k <= 3):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i + disp_sym[i,j]) < height):
#                             new_sub[i + disp_sym[i,j], j] += dif[i,j]
#                             
#             elif(k >= 4 and k <= 7):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j + disp_sym[i,j]) < width):
#                             new_sub[i, j + disp_sym[i,j]] += dif[i,j]
#                             
#             elif( k >= 9 and k <= 12):
#                 for i in range(height):
#                     for j in range(width):
#                         if((j - disp_sym[i,j]) > 0):
#                             new_sub[i, j - disp_sym[i,j]] += dif[i,j]
#                             
#             elif( k >= 13 and k <= 16):
#                 for i in range(height):
#                     for j in range(width):
#                         if((i - disp_sym[i,j]) > 0):
#                             new_sub[i - disp_sym[i,j], j] += dif[i,j]                
#             
#             new_sub = np.clip(new_sub, 0, 255)
#             new_sub = new_sub.astype(np.uint8)
#             
#             SSIM_R = SSIM(sub[:,:,2], sub_re[:,:,2])
#             SSIM_G = SSIM(sub[:,:,1], sub_re[:,:,1])
#             SSIM_B = SSIM(sub[:,:,0], sub_re[:,:,0])
#             SSIM_sub_re = (SSIM_R+SSIM_B+SSIM_G)/3
#             
#             
#             SSIM_R = SSIM(sub[:,:,2], new_sub[:,:,2])
#             SSIM_G = SSIM(sub[:,:,1], new_sub[:,:,1])
#             SSIM_B = SSIM(sub[:,:,0], new_sub[:,:,0])
#             SSIM_new_sub = (SSIM_R+SSIM_B+SSIM_G)/3
#             
#             ws.append([str(k), cv2.PSNR(sub,sub_re), cv2.PSNR(sub, new_sub), SSIM_sub_re, SSIM_new_sub])
#             print(str(k))
#             print("sub_re PSNR: " + str(cv2.PSNR(sub, sub_re)))
#             print("new_sub PSNR: " + str(cv2.PSNR(sub, new_sub)))
#             print("SSIM sub_re: " + str(SSIM_sub_re))
#             print("SSIM new_sub: " + str(SSIM_new_sub))
#             cv2.imwrite("warping_dataset/bike/after_compensate/"+ str(k) +"_symmetric_withoutOcclusion_twoside_thresh1.png", new_sub) 
#     
#     wb.save('warping_dataset/bike/symmetric_compensate.xlsx') 
#     #%% blend自己那圈的disparity map來做warping
#     print("Warping by blended disparity maps on the same ring...")
#     wb = Workbook()
#     ws = wb.active
#     ws.append(["","bilinear_psnr","compensate_ring_psnr","bilinear_SSIM","compensate_ring_SSIM"])
#     
#     disp_arr_ring = np.array([disp_arr_mid[3],disp_arr_mid[7],disp_arr_mid[9],disp_arr_mid[13],disp_arr_mid[17],disp_arr_mid[18],disp_arr_mid[19],disp_arr_mid[20]])
#     disp_ring = np.zeros((height,width))
#     for i in range(height):
#         for j in range(width):
#             disp_ring[i,j] = np.median(disp_arr_ring[:,i,j])
#     disp_ring = disp_ring.astype(np.uint8)
#     
#     test = [3,7,9,13]
#     for k in test:
#         print("Warping difference by blended ring disparity map on " + str(k) + "...")
#         sub = cv2.imread("warping_dataset/bike/HD_data/"+ str(k) +".png")
#         sub_down = cv2.resize(sub, (int(width*scale_percent), int(height*scale_percent)), interpolation = cv2.INTER_LINEAR)#換圖要改
#         sub_re = cv2.resize(sub_down, (width, height), interpolation = cv2.INTER_LINEAR)#換圖要改
#         new_sub = sub_re.astype(np.int32)
#         if(k == 3):
#             for i in range(height):
#                 for j in range(width):
#                      if((i + disp_ring[i,j]) < height):
#                          new_sub[i + disp_ring[i,j], j] += dif[i,j]
#         if(k == 7):
#             for i in range(height):
#                 for j in range(width):
#                     if((j + disp_ring[i,j]) < width):
#                         new_sub[i, j + disp_ring[i,j]] += dif[i,j]                    
#         if(k == 9):
#             for i in range(height):
#                 for j in range(width):
#                     if((j - disp_ring[i,j]) > 0):
#                         new_sub[i, j - disp_ring[i,j]] += dif[i,j]                    
#         if(k == 13):
#             for i in range(height):
#                 for j in range(width):
#                     if((i - disp_ring[i,j]) > 0):
#                         new_sub[i - disp_ring[i,j], j] += dif[i,j]   
#                         
#         new_sub = np.clip(new_sub, 0, 255)
#         new_sub = new_sub.astype(np.uint8)
#         
#         SSIM_R = SSIM(sub[:,:,2], sub_re[:,:,2])
#         SSIM_G = SSIM(sub[:,:,1], sub_re[:,:,1])
#         SSIM_B = SSIM(sub[:,:,0], sub_re[:,:,0])
#         SSIM_sub_re = (SSIM_R+SSIM_B+SSIM_G)/3
#         
#         SSIM_R = SSIM(sub[:,:,2], new_sub[:,:,2])
#         SSIM_G = SSIM(sub[:,:,1], new_sub[:,:,1])
#         SSIM_B = SSIM(sub[:,:,0], new_sub[:,:,0])
#         SSIM_new_sub = (SSIM_R+SSIM_B+SSIM_G)/3
#         
#         ws.append([str(k), cv2.PSNR(sub,sub_re), cv2.PSNR(sub, new_sub), SSIM_sub_re, SSIM_new_sub])
#         print(str(k))
#         print("sub_re PSNR: " + str(cv2.PSNR(sub, sub_re)))
#         print("new_sub PSNR: " + str(cv2.PSNR(sub, new_sub)))
#         print("SSIM sub_re: " + str(SSIM_sub_re))
#         print("SSIM new_sub: " + str(SSIM_new_sub))
#         cv2.imwrite("warping_dataset/bike/after_compensate/"+ str(k) +"_after_ring_compensate.png", new_sub)    
#                       
#     wb.save('warping_dataset/bike/ring_compensate.xlsx') 
# =============================================================================

